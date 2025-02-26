from django.contrib import admin
from django.http import HttpResponse
import csv
import openpyxl
from openpyxl.styles import Alignment

from .models import User, TestNSI, TestResult, Attempt


@admin.register(Attempt)
class AttemptAdmin(admin.ModelAdmin):
    list_display = ('user_id', 'test_id', 'try_count')


@admin.register(User)
class UserAdmin(admin.ModelAdmin):
    list_display = ('username', 'age', 'education', 'speciality')
    actions = ['export_as_csv', 'export_as_xlsx']

    def export_as_csv(self, request, queryset):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=users_test_results.csv'
        writer = csv.writer(response)

        # Заголовки столбцов
        headers = [
            'ИД', 'Возраст', 'Образование', 'Специальность', 'Место жительства', 'Рост', 'Вес', 'Ведущая рука',
            'Заболевания', 'Курение', 'Алкоголь', 'Спорт', 'Бессонница', 'Текущее самочувствие', 'Геймер'
        ]
        # Добавляем заголовки для каждого теста
        tests = TestNSI.objects.all()
        for test in tests:
            headers.extend([f'{test.test_name} - {field.name}' for field in TestResult._meta.fields if field.name not in ['id', 'test', 'user']])
        writer.writerow(headers)

        # Данные
        for user in queryset:
            user_data = [
                user.id, user.age, user.education, user.speciality, user.residence, user.height, user.weight,
                user.lead_hand, user.diseases, user.smoking, user.alcohol, user.sport, user.insomnia,
                user.current_health, user.gaming
            ]
            for test in tests:
                results = TestResult.objects.filter(user=user, test=test)
                if results.exists():
                    for result in results:
                        result_data = [getattr(result, field.name) for field in TestResult._meta.fields if field.name not in ['id', 'test', 'user']]
                        writer.writerow(user_data + result_data)
                else:
                    # Если нет результатов, заполняем пустыми значениями
                    writer.writerow(user_data + [''] * len([field for field in TestResult._meta.fields if field.name not in ['id', 'test', 'user']]))
        return response

    def export_as_xlsx(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=users_test_results.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active

        # Заголовки столбцов для информации о пользователе
        user_headers = [
            'ИД', 'Возраст', 'Образование', 'Специальность', 'Место жительства', 'Рост', 'Вес', 'Ведущая рука',
            'Заболевания', 'Курение', 'Алкоголь', 'Спорт', 'Бессонница', 'Текущее самочувствие', 'Геймер'
        ]
        ws.append(user_headers)

        # Получаем все тесты
        tests = TestNSI.objects.all()

        # Словарь для хранения информации о столбцах, относящихся к каждому тесту
        test_columns = {}

        # Текущий индекс столбца (начинаем с количества столбцов пользователя)
        current_col_idx = len(user_headers) + 1

        # Добавляем заголовки для каждого теста
        for test in tests:
            # Название теста
            test_name = test.test_name
            # Поля теста (исключаем 'id', 'test', 'user')
            test_fields = [field.name for field in TestResult._meta.fields if field.name not in ['id', 'test', 'user']]
            # Количество столбцов для этого теста
            num_columns = len(test_fields)
            # Записываем название теста в объединённую ячейку
            ws.merge_cells(start_row=1, start_column=current_col_idx, end_row=1,
                           end_column=current_col_idx + num_columns - 1)
            ws.cell(row=1, column=current_col_idx, value=test_name)
            # Записываем названия полей теста
            for i, field in enumerate(test_fields):
                ws.cell(row=2, column=current_col_idx + i, value=field)
            # Сохраняем информацию о столбцах для этого теста
            test_columns[test.test_id] = {
                'start_col': current_col_idx,
                'end_col': current_col_idx + num_columns - 1,
                'fields': test_fields
            }
            # Обновляем текущий индекс столбца
            current_col_idx += num_columns

        # Данные
        row_idx = 3  # Начинаем с третьей строки (после заголовков)
        for user in queryset:
            user_data = [
                user.id, user.age, user.education, user.speciality, user.residence, user.height, user.weight,
                user.lead_hand, user.diseases, user.smoking, user.alcohol, user.sport, user.insomnia,
                user.current_health, user.gaming
            ]
            # Считаем общее количество попыток для пользователя
            total_attempts = sum(Attempt.objects.filter(user=user).values_list('try_count', flat=True))
            if total_attempts == 0:
                total_attempts = 1  # Минимум одна строка, даже если попыток нет

            # Записываем данные пользователя
            for col_idx, value in enumerate(user_data, start=1):
                # Объединяем ячейки для информации о пользователе (столбцы 1-15)
                ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx + total_attempts - 1,
                               end_column=col_idx)
                ws.cell(row=row_idx, column=col_idx, value=value)

            # Записываем результаты тестов
            for test in tests:
                results = TestResult.objects.filter(user=user, test=test)
                if results.exists():
                    for result in results:
                        # Записываем результаты теста
                        test_info = test_columns[test.test_id]
                        for i, field in enumerate(test_info['fields']):
                            value = getattr(result, field)
                            # Убираем временную зону, если это datetime
                            if field == 'complete_time' and value is not None:
                                value = value.replace(tzinfo=None)
                            ws.cell(row=row_idx, column=test_info['start_col'] + i, value=value)
                        # Переходим на следующую строку для следующей попытки
                        row_idx += 1
                else:
                    # Если нет результатов, оставляем пустые ячейки
                    for i in range(len(test_columns[test.test_id]['fields'])):
                        ws.cell(row=row_idx, column=test_columns[test.test_id]['start_col'] + i, value='')
                    # Переходим на следующую строку
                    row_idx += 1

        wb.save(response)
        return response

    export_as_csv.short_description = "Export Selected as CSV"
    export_as_xlsx.short_description = "Export Selected as XLSX"


@admin.register(TestNSI)
class TestNSIAdmin(admin.ModelAdmin):
    list_display = ('test_name', 'title_all', 'title_correct')
    actions = ['export_as_csv', 'export_as_xlsx']

    def export_as_csv(self, request, queryset):
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename={meta}.csv'
        writer = csv.writer(response)
        writer.writerow(field_names)
        for obj in queryset:
            writer.writerow([getattr(obj, field) for field in field_names])
        return response

    def export_as_xlsx(self, request, queryset):
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(field_names)
        for obj in queryset:
            ws.append([getattr(obj, field) for field in field_names])
        wb.save(response)
        return response

    export_as_csv.short_description = "Export Selected as CSV"
    export_as_xlsx.short_description = "Export Selected as XLSX"


@admin.register(TestResult)
class TestResultAdmin(admin.ModelAdmin):
    list_display = ('user', 'test', 'try_number', 'accuracy')
    actions = ['export_as_csv', 'export_as_xlsx']

    def export_as_csv(self, request, queryset):
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename={meta}.csv'
        writer = csv.writer(response)
        writer.writerow(field_names)
        for obj in queryset:
            writer.writerow([getattr(obj, field) for field in field_names])
        return response

    def export_as_xlsx(self, request, queryset):
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={meta}.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(field_names)
        for obj in queryset:
            ws.append([getattr(obj, field) for field in field_names])
        wb.save(response)
        return response

    export_as_csv.short_description = "Export Selected as CSV"
    export_as_xlsx.short_description = "Export Selected as XLSX"
